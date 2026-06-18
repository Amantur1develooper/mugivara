#!/usr/bin/env python
"""
Импорт меню ресторана "Булак" из Excel файла.

Запуск (из корня проекта):
    python import_bulak_menu.py

Параметры:
    --file PATH            путь к .xlsx (по умолчанию ищет "Меню Булак*.xlsx" рядом)
    --dry-run              показать что будет импортировано, без записи в БД
    --restaurant-id N      использовать существующий ресторан (не создавать)
    --branch-id N          использовать существующий филиал (не создавать)
"""

import argparse
import glob
import os
import re
import sys
from decimal import Decimal

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")

import django
django.setup()

from django.db import transaction
from core.models import Restaurant, Branch
from catalog.models import (
    MenuSet, Category, Item, ItemCategory,
    BranchMenuSet, BranchCategory, BranchItem, BranchCategoryItem,
)

# ── Пометки «убрать» в столбце 9 ─────────────────────────────────────────────
SKIP_MARKS = {"убираем", "убрать", "убираем ", "убрать "}

# ── Секции, чьи подразделы получают префикс (напр. "Детское: Закуски") ────────
PREFIX_SECTIONS = {"Детское"}


# ── Парсинг Excel ─────────────────────────────────────────────────────────────

def _clean(v) -> str:
    """Привести значение ячейки к чистой строке."""
    if v is None:
        return ""
    return " ".join(str(v).split())  # убираем лишние пробелы и переносы


def _name_from_row(col0, col1) -> str:
    """
    Извлекаем название позиции.
    Обычно это col1. Если col1 пустой, пробуем col0 (бывает '10Удон с говядиной').
    """
    name = _clean(col1)
    if name:
        return name
    # col0 может быть '10Удон с говядиной' или '1+K107...' — берём только текстовую часть
    raw = _clean(col0)
    m = re.match(r"^\d+(.*)", raw)
    return m.group(1).strip() if m and m.group(1).strip() else raw


def parse_excel(path: str) -> list[dict]:
    """
    Читает лист 'МЕНЮ' и возвращает список:
      [{"category": str, "name": str, "description": str, "price": int}, ...]
    """
    try:
        import openpyxl
    except ImportError:
        print("Установите openpyxl:  pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    if "МЕНЮ" not in wb.sheetnames:
        print(f"Лист 'МЕНЮ' не найден в {path}. Листы: {wb.sheetnames}")
        sys.exit(1)

    ws = wb["МЕНЮ"]

    current_section  = None   # суперсекция (Детское, Вторые …)
    current_category = None   # текущая категория для позиций

    items = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        col0 = row[0] if len(row) > 0 else None
        col1 = row[1] if len(row) > 1 else None
        col2 = row[2] if len(row) > 2 else None
        col8 = row[8] if len(row) > 8 else None
        col9 = row[9] if len(row) > 9 else None

        # Пустая строка
        name_raw = col1 or col0
        if not name_raw or not _clean(name_raw):
            continue

        # Помечено «убрать»
        if col9 and _clean(col9).lower() in SKIP_MARKS:
            continue

        # Есть ли цена?
        has_price = isinstance(col8, (int, float)) and col8 > 0

        # ── Заголовок категории ──
        if not has_price:
            header = _clean(col1 or "")
            if not header:
                continue

            if header in PREFIX_SECTIONS:
                # Запомним суперсекцию, но не меняем current_category
                # (первая подкатегория внутри неё станет реальной категорией)
                current_section = header
            elif current_section in PREFIX_SECTIONS:
                # Мы внутри суперсекции — добавляем префикс
                current_category = f"{current_section}: {header}"
            else:
                current_section  = None
                current_category = header
            continue

        # ── Позиция меню ──
        if current_category is None:
            continue  # до первого заголовка — пропускаем

        name = _name_from_row(col0, col1)
        if not name:
            continue

        desc = _clean(col2)
        # Убираем случайные пробелы и неразрывные пробелы
        desc = desc.replace(" ", " ").strip()

        items.append({
            "category":    current_category,
            "name":        name,
            "description": desc,
            "price":       int(col8),
        })

    return items


# ── Импорт в БД ───────────────────────────────────────────────────────────────

def import_data(
    items: list[dict],
    *,
    dry_run: bool = False,
    restaurant_id: int | None = None,
    branch_id: int | None = None,
):
    # Собираем статистику по категориям
    cats: dict[str, list] = {}
    for it in items:
        cats.setdefault(it["category"], []).append(it)

    total_cats  = len(cats)
    total_items = len(items)
    items_with_desc = sum(1 for i in items if i["description"])

    print(f"\nКатегорий: {total_cats}")
    print(f"Позиций:   {total_items}  (с описанием: {items_with_desc})")

    if dry_run:
        print("\n[DRY RUN] — в БД ничего не пишется.\n")
        for cat, cat_items in cats.items():
            print(f"  [{cat}]  ({len(cat_items)} позиций)")
            for i in cat_items:
                desc = f"  — {i['description'][:60]}…" if i["description"] else ""
                print(f"    • {i['name']}  {i['price']} сом{desc}")
        return

    with transaction.atomic():
        # ── Ресторан ──
        if restaurant_id:
            restaurant = Restaurant.objects.get(pk=restaurant_id)
            print(f"\nРесторан: {restaurant} (id={restaurant.pk})")
        else:
            restaurant, cr = Restaurant.objects.get_or_create(
                slug="bulak",
                defaults={
                    "name_ru": "Булак",
                    "name_ky": "Булак",
                    "name_en": "Bulak",
                    "is_active": True,
                },
            )
            print(f"\n{'Создан' if cr else 'Найден'} ресторан: {restaurant} (id={restaurant.pk})")

        # ── Филиал ──
        if branch_id:
            branch = Branch.objects.get(pk=branch_id)
            print(f"Филиал: {branch} (id={branch.pk})")
        else:
            branch, cr = Branch.objects.get_or_create(
                restaurant=restaurant,
                name_ru="Булак — Основной",
                defaults={
                    "is_active": True,
                    "delivery_enabled": False,
                    "pay_cash_enabled": True,
                    "pay_online_enabled": True,
                },
            )
            print(f"{'Создан' if cr else 'Найден'} филиал: {branch} (id={branch.pk})")

        # ── MenuSet ──
        menu_set, cr = MenuSet.objects.get_or_create(
            restaurant=restaurant,
            name="Основное меню",
            defaults={"is_active": True},
        )
        print(f"{'Создан' if cr else 'Найден'} MenuSet: {menu_set} (id={menu_set.pk})")
        BranchMenuSet.objects.get_or_create(
            branch=branch, menu_set=menu_set,
            defaults={"is_active": True},
        )

        # ── Категории и позиции ──
        stat = dict(cats_new=0, items_new=0, items_skipped=0)

        for sort_cat, (cat_name, cat_items) in enumerate(cats.items(), start=1):
            category, cr = Category.objects.get_or_create(
                menu_set=menu_set,
                name_ru=cat_name,
                defaults={"name_ky": cat_name, "name_en": cat_name},
            )
            if cr:
                stat["cats_new"] += 1

            branch_cat, _ = BranchCategory.objects.get_or_create(
                branch=branch,
                category=category,
                defaults={"sort_order": sort_cat, "is_active": True},
            )

            for sort_item, it in enumerate(cat_items, start=1):
                price = Decimal(it["price"])

                item, item_cr = Item.objects.get_or_create(
                    restaurant=restaurant,
                    name_ru=it["name"],
                    defaults={
                        "name_ky":        it["name"],
                        "name_en":        it["name"],
                        "base_price":     price,
                        "description_ru": it["description"],
                    },
                )

                # Обновляем описание у существующих позиций если оно было пустым
                if not item_cr and it["description"] and not item.description_ru:
                    item.description_ru = it["description"]
                    item.save(update_fields=["description_ru"])

                if item_cr:
                    stat["items_new"] += 1
                else:
                    stat["items_skipped"] += 1

                ItemCategory.objects.get_or_create(
                    item=item, category=category,
                    defaults={"sort_order": sort_item},
                )

                branch_item, _ = BranchItem.objects.get_or_create(
                    branch=branch, item=item,
                    defaults={
                        "price":              price,
                        "is_available":       True,
                        "sort_order":         sort_item,
                        "delivery_available": True,
                    },
                )

                BranchCategoryItem.objects.get_or_create(
                    branch_category=branch_cat,
                    branch_item=branch_item,
                    defaults={"sort_order": sort_item},
                )

    print(f"\n✅ Готово!")
    print(f"   Категорий создано:  {stat['cats_new']}")
    print(f"   Позиций создано:    {stat['items_new']}")
    print(f"   Позиций пропущено:  {stat['items_skipped']} (уже существуют)")
    print(f"\n   Ресторан ID: {restaurant.pk}")
    print(f"   Филиал ID:   {branch.pk}")
    print(f"   MenuSet ID:  {menu_set.pk}")


# ── Точка входа ───────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Импорт меню Булак из Excel")
    ap.add_argument("--file",          default=None,  help="Путь к .xlsx файлу")
    ap.add_argument("--dry-run",       action="store_true")
    ap.add_argument("--restaurant-id", type=int, default=None)
    ap.add_argument("--branch-id",     type=int, default=None)
    args = ap.parse_args()

    # Найти файл автоматически если не указан
    xlsx_path = args.file
    if not xlsx_path:
        candidates = glob.glob(os.path.join(BASE_DIR, "Меню Булак*.xlsx"))
        if not candidates:
            print("Файл не найден. Укажите путь: --file /path/to/menu.xlsx")
            sys.exit(1)
        xlsx_path = candidates[0]

    print(f"Файл: {xlsx_path}")

    items = parse_excel(xlsx_path)
    if not items:
        print("Позиции не найдены — проверьте файл.")
        sys.exit(1)

    import_data(
        items,
        dry_run=args.dry_run,
        restaurant_id=args.restaurant_id,
        branch_id=args.branch_id,
    )


if __name__ == "__main__":
    main()
