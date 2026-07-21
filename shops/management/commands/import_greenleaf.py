"""
Импорт товаров из catalog.xlsx в магазин «СЦ Гринлиф Самат».

Запуск:
    python manage.py import_greenleaf
    python manage.py import_greenleaf --dry-run
    python manage.py import_greenleaf --store-id 3 --branch-id 5
    python manage.py import_greenleaf --file /path/to/catalog.xlsx
"""

from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand

from shops.models import Store, StoreBranch, StoreCategory, StoreProduct, StoreStock

DEFAULT_XLSX = Path(__file__).resolve().parents[3] / "catalog.xlsx"
STORE_NAME   = "СЦ Гринлиф Самат"
STORE_SLUG   = "sc-greenleaf-samat"
BRANCH_NAME  = "СЦ Гринлиф Самат — Бишкек"


class Command(BaseCommand):
    help = "Импортирует товары из catalog.xlsx в магазин «СЦ Гринлиф Самат»"

    def add_arguments(self, parser):
        parser.add_argument("--file", default=str(DEFAULT_XLSX),
                            help="Путь к Excel-файлу (по умолчанию: catalog.xlsx рядом с manage.py)")
        parser.add_argument("--store-id", type=int, default=None,
                            help="ID существующего Store (если не указан — будет создан)")
        parser.add_argument("--branch-id", type=int, default=None,
                            help="ID существующего StoreBranch (если не указан — будет создан)")
        parser.add_argument("--dry-run", action="store_true",
                            help="Только вывод, без записи в БД")

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write(self.style.ERROR("openpyxl не установлен. Выполните: pip install openpyxl"))
            return

        dry = options["dry_run"]
        xlsx_path = Path(options["file"])

        if not xlsx_path.exists():
            self.stderr.write(self.style.ERROR(f"Файл не найден: {xlsx_path}"))
            return

        # ── 1. Магазин ──────────────────────────────────────────────────────
        if options["store_id"]:
            store = Store.objects.get(pk=options["store_id"])
            self.stdout.write(f"Используется существующий магазин: {store}")
        elif not dry:
            store, created = Store.objects.get_or_create(
                slug=STORE_SLUG,
                defaults={"name_ru": STORE_NAME, "is_active": True},
            )
            self.stdout.write(f"{'Создан' if created else 'Найден'} магазин: {store.name_ru} (id={store.pk})")
        else:
            store = None

        # ── 2. Филиал ───────────────────────────────────────────────────────
        if options["branch_id"]:
            branch = StoreBranch.objects.get(pk=options["branch_id"])
            self.stdout.write(f"Используется существующий филиал: {branch}")
        elif not dry and store:
            branch, created = StoreBranch.objects.get_or_create(
                store=store,
                name_ru=BRANCH_NAME,
                defaults={
                    "city": StoreBranch.City.BISHKEK,
                    "is_active": True,
                },
            )
            self.stdout.write(f"{'Создан' if created else 'Найден'} филиал: {branch.name_ru} (id={branch.pk})")
        else:
            branch = None

        # ── 3. Чтение Excel ─────────────────────────────────────────────────
        self.stdout.write(f"Читаю файл: {xlsx_path}")
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        # Ищем лист «Каталог», иначе берём первый
        sheet_name = "Каталог" if "Каталог" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        self.stdout.write(f"Лист: «{sheet_name}»")

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            self.stderr.write(self.style.ERROR("Файл пустой"))
            return

        # Пропускаем заголовок
        data_rows = rows[1:]
        self.stdout.write(f"Строк данных: {len(data_rows)}")

        # Единственная категория — все товары без разбивки по категориям в файле
        category_cache: dict[str, StoreCategory] = {}

        def get_category(name: str) -> StoreCategory | None:
            if dry or store is None:
                return None
            if name not in category_cache:
                cat, _ = StoreCategory.objects.get_or_create(
                    store=store,
                    name_ru=name,
                    defaults={"sort_order": len(category_cache) * 10},
                )
                category_cache[name] = cat
            return category_cache[name]

        imported = updated = skipped = 0

        for i, row in enumerate(data_rows, start=2):
            # Колонки: Код | Артикул | Наименование | Кол-во в коробке | Доступно | Цена 1 | Цена 2 | Баллы
            if len(row) < 6:
                skipped += 1
                continue

            barcode  = str(row[1]).strip() if row[1] is not None else ""
            name     = str(row[2]).strip() if row[2] is not None else ""
            qty_raw  = row[4]
            price_raw = row[5]

            if not name:
                skipped += 1
                continue

            # Цена
            try:
                price = Decimal(str(price_raw)).quantize(Decimal("0.01"))
            except (InvalidOperation, TypeError):
                # Нет цены — пропускаем
                self.stdout.write(self.style.WARNING(f"  [строка {i}] Нет цены, пропускаю: {name}"))
                skipped += 1
                continue

            # Остаток
            try:
                qty = Decimal(str(qty_raw)).quantize(Decimal("1")) if qty_raw is not None else Decimal("0")
            except (InvalidOperation, TypeError):
                qty = Decimal("0")

            self.stdout.write(
                f"  [строка {i}] {name[:55]:55s} | арт: {barcode:<12} | {price:>8} сом | склад: {qty}"
            )

            if dry:
                continue

            cat = get_category("Основная категория")

            # Товар — уникальный ключ: barcode + store
            if barcode:
                product, created = StoreProduct.objects.get_or_create(
                    store=store,
                    barcode=barcode,
                    defaults={
                        "name_ru": name,
                        "category": cat,
                        "price": price,
                        "is_active": True,
                    },
                )
                if not created:
                    product.name_ru = name
                    product.price   = price
                    if cat:
                        product.category = cat
                    product.save(update_fields=["name_ru", "price", "category"])
                    updated += 1
                else:
                    imported += 1
            else:
                # Нет артикула — ищем по названию
                product, created = StoreProduct.objects.get_or_create(
                    store=store,
                    name_ru=name,
                    defaults={
                        "category": cat,
                        "price": price,
                        "is_active": True,
                    },
                )
                if not created:
                    product.price = price
                    if cat:
                        product.category = cat
                    product.save(update_fields=["price", "category"])
                    updated += 1
                else:
                    imported += 1

            # Остаток
            if branch:
                stock, _ = StoreStock.objects.get_or_create(
                    branch=branch,
                    product=product,
                    defaults={"qty": qty},
                )
                if not _:
                    stock.qty = qty
                    stock.save(update_fields=["qty"])

        self.stdout.write(self.style.SUCCESS(
            f"\nГотово! Создано: {imported}, обновлено: {updated}, пропущено: {skipped}"
        ))
        if dry:
            self.stdout.write(self.style.WARNING("(dry-run — ничего не сохранено в БД)"))
