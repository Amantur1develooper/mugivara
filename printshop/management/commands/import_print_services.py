"""
Импорт полиграфического центра и его услуг из Excel-файла.

Использование:
    python manage.py import_print_services                     # файл по умолчанию
    python manage.py import_print_services --file=другой.xlsx  # свой файл
    python manage.py import_print_services --center="Моя типография" --phone="+996700123456"
    python manage.py import_print_services --reset             # удалить старый центр и создать заново
"""

import re
import os
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError

DEFAULT_XLSX = os.path.join(
    os.path.dirname(__file__),           # commands/
    "..", "..", "..",                    # sanzhi/
    "список услуг для полиграфии.xlsx",
)


def _parse_price(raw) -> Decimal:
    """Извлекает первое число из строки вида '6 сом', '1 600 сом', '600/от 10шт 400'."""
    if raw is None:
        return Decimal("0")
    # убираем пробелы внутри числа (1 600 → 1600)
    s = re.sub(r"(\d)\s+(\d)", r"\1\2", str(raw))
    m = re.search(r"\d+(?:[.,]\d+)?", s)
    if not m:
        return Decimal("0")
    try:
        return Decimal(m.group().replace(",", "."))
    except InvalidOperation:
        return Decimal("0")


def _cell(row, idx, default=""):
    """Безопасно достаёт ячейку по индексу и возвращает stripped строку."""
    try:
        val = row[idx]
    except IndexError:
        return default
    if val is None:
        return default
    return str(val).strip()


class Command(BaseCommand):
    help = "Импортирует полиграфический центр и услуги из Excel-файла"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file", default=None,
            help="Путь к xlsx-файлу (по умолчанию: список услуг для полиграфии.xlsx в корне проекта)",
        )
        parser.add_argument(
            "--center", default="Полиграфия Санжи",
            help="Название полиграфического центра",
        )
        parser.add_argument(
            "--branch", default="Основной филиал",
            help="Название филиала",
        )
        parser.add_argument(
            "--phone", default="",
            help="Телефон филиала (например +996700123456)",
        )
        parser.add_argument(
            "--whatsapp", default="",
            help="WhatsApp номер (цифры, без +)",
        )
        parser.add_argument(
            "--address", default="",
            help="Адрес филиала",
        )
        parser.add_argument(
            "--slug", default="sanzhi-print",
            help="Slug для PrintCenter (уникальный, латиница)",
        )
        parser.add_argument(
            "--reset", action="store_true",
            help="Удалить существующий центр с таким slug и создать заново",
        )

    # ── helpers ──────────────────────────────────────────────────────────────

    def _ok(self, msg):
        self.stdout.write(self.style.SUCCESS(f"  ✓ {msg}"))

    def _info(self, msg):
        self.stdout.write(f"  → {msg}")

    def _warn(self, msg):
        self.stdout.write(self.style.WARNING(f"  ⚠ {msg}"))

    # ── main ─────────────────────────────────────────────────────────────────

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("Установите openpyxl: pip install openpyxl")

        # resolve file path
        xlsx_path = options["file"] or os.path.normpath(DEFAULT_XLSX)
        if not os.path.exists(xlsx_path):
            raise CommandError(f"Файл не найден: {xlsx_path}")

        from printshop.models import PrintCenter, PrintBranch, PrintCategory, PrintProduct

        # ── 1. PrintCenter ──────────────────────────────────────────────────
        self.stdout.write(self.style.MIGRATE_HEADING("\n=== Полиграфический центр ==="))

        slug = options["slug"]

        if options["reset"]:
            deleted, _ = PrintCenter.objects.filter(slug=slug).delete()
            if deleted:
                self._warn(f"Удалён старый центр slug={slug}")

        center, created = PrintCenter.objects.get_or_create(
            slug=slug,
            defaults={"name_ru": options["center"], "is_active": True},
        )
        if created:
            self._ok(f"Создан центр: «{center.name_ru}»")
        else:
            self._info(f"Центр уже существует: «{center.name_ru}» (id={center.id})")

        # ── 2. PrintBranch ──────────────────────────────────────────────────
        self.stdout.write(self.style.MIGRATE_HEADING("\n=== Филиал ==="))

        branch, br_created = PrintBranch.objects.get_or_create(
            center=center,
            name_ru=options["branch"],
            defaults={
                "phone":    options["phone"],
                "whatsapp": options["whatsapp"],
                "address":  options["address"],
                "is_active": True,
            },
        )
        if br_created:
            self._ok(f"Создан филиал: «{branch.name_ru}»")
        else:
            self._info(f"Филиал уже существует: «{branch.name_ru}» (id={branch.id})")

        # ── 3. Read Excel ───────────────────────────────────────────────────
        self.stdout.write(self.style.MIGRATE_HEADING("\n=== Импорт категорий и услуг ==="))

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

        total_cats = 0
        total_prods = 0
        total_skipped = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            cat_name = sheet_name.strip()

            # Determine column layout by scanning header row (row 2)
            # Layout A: (None, name, description, price, category)  — "бумажный"
            # Layout B: (None, None, name, description, price)       — остальные
            #
            # We detect by checking if row[1] (col B) looks like a header word
            header_row = None
            name_col = desc_col = price_col = None

            for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                r = [str(c).strip().lower() if c else "" for c in row]
                if any(w in " ".join(r) for w in ("название", "наименование")):
                    header_row = r
                    break

            if header_row is None:
                self._warn(f"Лист «{cat_name}»: не найдена строка заголовков — пропускаем")
                continue

            for i, cell in enumerate(header_row):
                if cell in ("название услуги", "название"):
                    name_col = i
                elif "наименование" in cell:
                    name_col = i
                if "описание" in cell:
                    desc_col = i
                if "цена" in cell:
                    price_col = i

            if name_col is None:
                self._warn(f"Лист «{cat_name}»: не найдена колонка с названием — пропускаем")
                continue

            # Get or create category
            cat, cat_new = PrintCategory.objects.get_or_create(
                center=center,
                name_ru=cat_name,
                defaults={"is_active": True, "sort_order": total_cats},
            )
            total_cats += 1 if cat_new else 0
            status = "создана" if cat_new else "уже есть"
            self._info(f"Категория «{cat_name}» — {status}")

            # Import products
            prod_count = 0
            for row in ws.iter_rows(min_row=3, values_only=True):
                name_raw = _cell(row, name_col)
                if not name_raw:
                    continue

                # skip if looks like a header repeat
                low = name_raw.lower()
                if any(w in low for w in ("наименование", "название", "услуги")):
                    continue

                desc_raw  = _cell(row, desc_col)  if desc_col  is not None else ""
                price_raw = _cell(row, price_col) if price_col is not None else ""
                price     = _parse_price(price_raw)

                # clean up newlines in description
                desc_clean = desc_raw.replace("\n", " ").strip()
                name_clean = name_raw.replace("\n", " ").strip()

                # idempotent: skip if product with same name in same category exists
                if PrintProduct.objects.filter(center=center, category=cat, name_ru=name_clean).exists():
                    total_skipped += 1
                    continue

                PrintProduct.objects.create(
                    center=center,
                    category=cat,
                    name_ru=name_clean,
                    description_ru=desc_clean,
                    base_price=price,
                    is_available=True,
                    sort_order=prod_count,
                )
                prod_count += 1
                total_prods += 1
                self.stdout.write(f"      + {name_clean[:50]:<50}  {price} сом")

        # ── Summary ─────────────────────────────────────────────────────────
        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(
            f"✅ Готово! Центр: «{center.name_ru}» | "
            f"Категорий создано: {total_cats} | "
            f"Услуг добавлено: {total_prods} | "
            f"Уже существовало: {total_skipped}"
        ))
        self.stdout.write(
            f"   🔗 Admin: /admin/printshop/printcenter/{center.id}/change/"
        )
